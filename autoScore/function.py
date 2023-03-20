import subprocess
from pathlib import Path
from function import *
import natsort # pip install natsort -> if not used: remove natsort.natsorted
import re
import openpyxl
import ast
import astunparse
import sys
import os
def getSubprocessesInputPrompt(pythonFilePath):
    with open(pythonFilePath, 'r' ,encoding="utf-8") as f:
        subprocess_code = f.read()
    allPrompts = []
    tree = None
    try:
        tree = ast.parse(subprocess_code)
    except SyntaxError:
        pass
    if tree is not None:
        for node in ast.walk(tree):
            if isinstance(node, ast.Call) and isinstance(node.func, ast.Name) and node.func.id == 'input':
                if len(node.args) > 0:
                    prompt_str = astunparse.unparse(node.args[0]).strip()
                    prompt = re.sub(r'^[\'\"]|[\'\"]$', '', prompt_str)
                else:
                    prompt = ''
                allPrompts.append(prompt)
                # print(repr(prompt))
                # return prompt
    return allPrompts

def getInputLines(inputFilePath):
    with open(inputFilePath, "r" , encoding='utf-8',errors='replace') as f:
        lines = f.readlines()
    return lines

def getAnswerLines(inputFilePath ):
    with open(inputFilePath, "r" , encoding='utf-8',errors='replace') as f:
        lines = f.readlines()
    return lines

def initLogFile(logFilePath):
    file_path = Path(logFilePath)

    # create or update the file
    with file_path.open(mode="w" , encoding='utf-8',errors='replace') as file:
        # delete the contents of the file
        file.truncate(0)

def initFailLogFile(logFilePath):
    file_path = Path(logFilePath)

    # create or update the file
    with file_path.open(mode="w" , encoding='utf-8',errors='replace') as file:
        # delete the contents of the file
        file.truncate(0)

def writeLogFile(logFilePath,logMsg):
    # define the path to the file
    file_path = Path(logFilePath)

    # open the file in append mode
    with file_path.open(mode="a" , encoding='utf-8',errors='replace') as file:
        # write data to the file
        file.write(logMsg)
def writeFailLogFile(logFilePath,logMsg):
    # define the path to the file
    file_path = Path(logFilePath)

    # open the file in append mode
    with file_path.open(mode="a" , encoding='utf-8',errors='replace') as file:
        # write data to the file
        file.write(logMsg)
def writeLogForNotFoundFile(logForNotFoundFile,NotFoundFileDirectory):
    # define the path to the file
    file_path = Path(logForNotFoundFile)

    # open the file in append mode
    with file_path.open(mode="a" , encoding='utf-8',errors='replace') as file:
        # write data to the file
        file.write(NotFoundFileDirectory)

def initLogForNotFoundFile(logForNotFoundFile):
    file_path = Path(logForNotFoundFile)

    # create or update the file
    with file_path.open(mode="w" , encoding='utf-8',errors='replace') as file:
        # delete the contents of the file
        file.truncate(0)


def initLogForExcelFile(logForExcelPath):
    file_path = Path(logForExcelPath)

    # create or update the file
    with file_path.open(mode="w" , encoding='utf-8',errors='replace') as file:
        # delete the contents of the file
        file.truncate(0)

def writeLogForExcelFile(logFilePath,pythonFilePathAndScore):
    # define the path to the file
    file_path = Path(logFilePath)

    # open the file in append mode
    with file_path.open(mode="a"  , encoding='utf-8',errors='replace') as file:
        # write data to the file
        file.write(pythonFilePathAndScore+"\n")

def executeHwFile(pythonFilePath,stdInput):
    command = [sys.executable,pythonFilePath]
    try:
        env = os.environ.copy()
        env['PYTHONIOENCODING']= 'utf-8'
        p = subprocess.Popen(command, env = env,stdin=subprocess.PIPE, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        # print(stdInput)
        output,err = p.communicate(input=stdInput.encode(),timeout=2)
        # output = p.communicate(input='5\n3\n'.encode())[0]
        # print(output.decode("utf-8"))
        # print(err.decode("utf-8"))

        p.kill()
        return output.decode("utf-8", 'replace'),err.decode("utf-8", 'replace')
    except Exception as expErr:
        # print(output.decode("utf-8"),exErr)
        return None,f"timeExpired or other error:\n {expErr}\n"

def scoreOnePythonFile_numeric(targetPythonFile):
    inputDir = Path("./input")
    answerDir = Path("./answer")

    allCorrect = 1
    logMsg = ""
    # logMsg = "---------------------------------------\n"
    # logMsg += targetPythonFile.parent.name+"/"+targetPythonFile.name+"\n"
    # print(targetPythonFile.parent.name.split("_")[0])
    # studentID=targetPythonFile.parent.name.split("_")[0]
    # print()
    for inputFilePath in natsort.natsorted(inputDir.glob('*.txt')):
        # if inputFilePath.endsw
        # print(inputFilePath)
        # print(inputFilePath.name)
        stdInput = "".join(getInputLines(inputFilePath))
        allPrompts = getSubprocessesInputPrompt(targetPythonFile)
        output,err = executeHwFile(targetPythonFile,stdInput)
        # print(allPrompts)
        # print(type(output))
        
        # print(err=="")
        if err == "":
            outputWithoutInputPrompt = output
            if output !=None and len(allPrompts)>0:
                for prompt in allPrompts:
                    outputWithoutInputPrompt = ''.join(re.split(prompt, output))
            print(output)
            print(outputWithoutInputPrompt)
            outputNumList = re.findall('[-+]?[\d]+[.]?[\d]*', outputWithoutInputPrompt)
            # print(outputNumList)

            answerFilePath = answerDir/inputFilePath.name
            answer = getAnswerLines(answerFilePath)
            # print(answer)
            answer = "".join(answer)
            answerNumList = re.findall('[-+]?[\d]+[.]?[\d]*', answer)
            # print(answerNumList)
            #---------------------------------------------------------------
            # check correctness
            correct = 1
            if len(outputNumList) != len(answerNumList):
                correct = 0
            else:
                for out,ans in zip(outputNumList,answerNumList):
                    if eval(out) != eval(ans):
                        correct = 0
            #---------------------------------------------------------------    
            if correct == 1:
                # print(f"{inputFilePath.name} pass")
                logMsg += "-------------------------------------\n"
                logMsg += f"test data {inputFilePath.name} pass\n"
                logMsg += "-------------------------------------\n"
                
            else :
                allCorrect = 0
                # print(f"{inputFilePath.name} something wrong")
                logMsg += "-------------------------------------\n"
                logMsg += f"test data {inputFilePath.name} something wrong\n"
                logMsg += "-------------------------------------\n"
                # print(f"output:")
                # logMsg += r"''''''''''''''''''''''''''''''''''''''" + "\n"
                # logMsg += "-------------------------------------------------------------\n"
                # logMsg += "-------------------------------------\n"
                logMsg += f"student output:\n"
                # logMsg += "-------------------------------------\n"
                logMsg += r"'''"+"\n"
                for out in outputNumList:
                    # print(out,end=" ",sep=" ")
                    logMsg += out + " "
                # print()
                # print(f"correct answer output:")
                logMsg += "\n"+r"'''"+"\n"
                # logMsg += "-------------------------------------\n"
                logMsg += f"correct answer output:\n"
                # logMsg += "-------------------------------------\n"
                logMsg += r"'''"+"\n"
                for out in answerNumList:
                    # print(out,end=" ",sep=" ")
                    logMsg += out + " "
                logMsg += "\n"+r"'''"+"\n"
                # logMsg += "-------------------------------------\n"
                logMsg += f"student original output (stdout):\n"
                # logMsg += "-------------------------------------\n"
                logMsg += r"'''"+"\n"
                logMsg += output
                logMsg += "\n"+r"'''"+"\n"
                # logMsg += "-------------------------------------------------------------\n"
                logMsg += "\n"
                # print("\n")
        else:
            logMsg += f"\nthere's errror in test data {inputFilePath.name}\n"
            logMsg += r"'''"+"\n"
            logMsg += err
            logMsg += r"'''"+"\n"
            # print(err)
            allCorrect = 0

    # print("\nfinal result:")
    # logMsg += "\nfinal result:\n"
    if allCorrect == 1:
        score = "100"
        # print("  all pass score : 100")
        # logMsg +="  all pass score : 100\n"
    else:
        # print("  fail")
        score = "fail"
        # logMsg += "  fail\n"
    
    # print(f"---------------log--------\n{logMsg}")

    return score,logMsg


def scoreOnePythonFile_graph(targetPythonFile):
    inputDir = Path("./input")
    answerDir = Path("./answer")

    allCorrect = 1
    logMsg = ""
    # logMsg = "---------------------------------------\n"
    # logMsg += targetPythonFile.parent.name+"/"+targetPythonFile.name+"\n"
    # print(targetPythonFile.parent.name.split("_")[0])
    # studentID=targetPythonFile.parent.name.split("_")[0]
    # print()
    for inputFilePath in natsort.natsorted(inputDir.glob('*.txt')):
        # print(inputFilePath)
        # print(inputFilePath.name)
        stdInput = "".join(getInputLines(inputFilePath))
        allPrompts = getSubprocessesInputPrompt(targetPythonFile)
        output,err = executeHwFile(targetPythonFile,stdInput)
        # print(allPrompts)
        # print(type(output))
        if output !=None:
            for prompt in allPrompts:
                output = ''.join(re.split(prompt, output))
        if err == "":
            # print(output)
            

            answerFilePath = answerDir/inputFilePath.name
            answer = getAnswerLines(answerFilePath)
            # print(answer)
            answer = "".join(answer)
            
            #---------------------------------------------------------------
            # check correctness
            correct = 1
            if answer not in output:
                correct = 0
            #---------------------------------------------------------------    
            if correct == 1:
                logMsg += "-------------------------------------\n"
                logMsg += f"test data {inputFilePath.name} pass\n"
                logMsg += "-------------------------------------\n"
            else :
                allCorrect = 0
                # print(f"{inputFilePath.name} something wrong")
                logMsg += "-------------------------------------\n"
                logMsg += f"test data {inputFilePath.name} something wrong\n"
                logMsg += "-------------------------------------\n"
                # print(f"output:")
                # logMsg += r"''''''''''''''''''''''''''''''''''''''" + "\n"
                # logMsg += "-------------------------------------------------------------\n"
                # logMsg += "-------------------------------------\n"
                
                # logMsg += "-------------------------------------\n"
                logMsg += f"correct answer output:\n"
                # logMsg += "-------------------------------------\n"
                logMsg += r"'''"+"\n"
                
                logMsg += answer
                
                logMsg += "\n"+r"'''"+"\n"
                # logMsg += "-------------------------------------\n"
                logMsg += f"student original output (stdout):\n"
                # logMsg += "-------------------------------------\n"
                logMsg += r"'''"+"\n"
                logMsg += output
                logMsg += "\n"+r"'''"+"\n"
                # logMsg += "-------------------------------------------------------------\n"
                logMsg += "\n"
                # print("\n")
        else:
            logMsg += f"\nthere's errror in test data {inputFilePath.name}\n"
            logMsg += r"'''"+"\n"
            logMsg += err
            logMsg += r"'''"+"\n"
            # print(err)
            allCorrect = 0

    # print("\nfinal result:")
    # logMsg += "\nfinal result:\n"
    if allCorrect == 1:
        score = "100"
        # print("  all pass score : 100")
        # logMsg +="  all pass score : 100\n"
    else:
        # print("  fail")
        score = "fail"
        # logMsg += "  fail\n"
    
    # print(f"---------------log--------\n{logMsg}")

    return score,logMsg




def writeToExcel(studentID,scoreOrErr):
    # Load the workbook
    workbook = openpyxl.load_workbook('statistic.xlsx')

    # Select the worksheet you want to work with
    worksheet = workbook.worksheets[0]

    # Define the value you're searching for in a specific column
    search_value = studentID

    # Define the column number where you want to write the data
    target_column = 2  # Column B

    foundOrNot = False
    # Loop through each row in the worksheet
    for row in range(1, worksheet.max_row + 1):
        # Check if the value in the target column matches the search value
        # print(worksheet.cell(row=row, column=1).value)
        if worksheet.cell(row=row, column=1).value!= None and worksheet.cell(row=row, column=1).value.lower() == search_value.lower():
            # Write the data to the next column in the same row
            worksheet.cell(row=row, column=target_column).value = scoreOrErr
            foundOrNot = True

    # Save the changes to the workbook
    workbook.save('statistic.xlsx')
    workbook.close()
    return foundOrNot